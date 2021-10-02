import openpyxl

# 1> Excel 作成
wb = openpyxl.Workbook()

# 2> Excel worksheet 作成
ws = wb.create_sheet('イカゲーム')

# 3> データ追加
ws['A1'] = '参加番号'
ws['B1'] = '名前'

ws['A2'] = 1
ws['B2'] = 'オ様'

# 4> Excel Save
wb.save(r'C:\Users\sumyo\Python_practice\Start Crawling\参加者_data.xlsx')