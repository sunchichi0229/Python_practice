import openpyxl

fpath = r'C:\Users\sumyo\Python_practice\Start Crawling\参加者_data.xlsx'

# 1>Open Excel
wb = openpyxl.load_workbook(fpath)

# 2>Choice ExcelSheet
ws = wb['イカゲーム']

# 3> Change data
ws['A3'] = 456
ws['B3'] = 'ソン様'

# 4> Save Excel
wb.save(fpath)